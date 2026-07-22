from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.colors import TwoSlopeNorm
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from skimage import measure

from .noise_filter import NoiseClassification
from .volumes import VolumeResult


@dataclass
class RasterTransform:
    origin_x: float
    origin_y: float
    pixel_size_m: float

    def pixel_to_geo(self, row: float, col: float) -> tuple[float, float]:
        x = self.origin_x + col * self.pixel_size_m
        y = self.origin_y - row * self.pixel_size_m
        return x, y


KIND_LABELS_RU = {"fill": "Насыпь/склад", "cut": "Выемка"}

def build_volume_statement(
    results: list[VolumeResult], transform: RasterTransform | None = None
) -> pd.DataFrame:
    rows = []
    for r in results:
        row = {
            "ID объекта": r.object_id,
            "Тип": KIND_LABELS_RU.get(r.kind, r.kind),
            "Площадь, м²": round(r.area_m2, 1),
            "Объём выемки, м³": round(r.cut_volume_m3, 1),
            "Объём насыпи, м³": round(r.fill_volume_m3, 1),
            "Чистый объём, м³": round(r.net_volume_m3, 1),
            "Средняя высота, м": round(r.mean_height_m, 2),
            "Макс. высота, м": round(r.max_height_m, 2),
            "Мин. высота, м": round(r.min_height_m, 2),
        }
        if transform is not None:
            x, y = transform.pixel_to_geo(r.centroid_row, r.centroid_col)
            row["X центроида"] = round(x, 2)
            row["Y центроида"] = round(y, 2)
        rows.append(row)
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values("ID объекта").reset_index(drop=True)
    return df


def build_noise_report(classifications: list[NoiseClassification]) -> pd.DataFrame:
    rows = [
        {
            "ID кандидата": c.object_id,
            "Площадь, м²": round(c.area_m2, 1),
            "Extent (заполненность bbox)": round(c.extent, 2),
            "Доля 'цвета техники'": f"{c.machinery_color_fraction:.0%}",
            "Результат": "включён в ведомость" if c.is_valid_object else "отсеян как шум",
            "Причина": c.reason,
        }
        for c in classifications
    ]
    return pd.DataFrame(rows)


def save_statement_xlsx(
    df: pd.DataFrame,
    path: str | Path,
    sheet_name: str = "Ведомость объёмов",
    extra_sheets: dict[str, pd.DataFrame] | None = None,
) -> None:
    path = Path(path)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        _write_formatted_sheet(writer, df, sheet_name, title="Ведомость объёмов по объектам учёта")
        if extra_sheets:
            for name, extra_df in extra_sheets.items():
                _write_formatted_sheet(writer, extra_df, name, title=name)


def _write_formatted_sheet(writer, df: pd.DataFrame, sheet_name: str, title: str) -> None:
    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
    ws = writer.sheets[sheet_name]

    n_cols = max(1, len(df.columns))
    ws.cell(row=1, column=1, value=title).font = Font(name="Arial", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(df.columns, start=1):
        values = [str(col_name)] + [str(v) for v in df[col_name]] if len(df) else [str(col_name)]
        max_len = max(len(v) for v in values)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(30, max_len + 3))

    for row_idx in range(3, 3 + len(df)):
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center")

    numeric_cols = [c for c in ("Объём выемки, м³", "Объём насыпи, м³", "Чистый объём, м³") if c in df.columns]
    if len(df) and numeric_cols:
        total_row = 3 + len(df) + 1
        ws.cell(row=total_row, column=1, value="ИТОГО").font = Font(name="Arial", bold=True)
        for col_name in numeric_cols:
            col_idx = list(df.columns).index(col_name) + 1
            cell = ws.cell(row=total_row, column=col_idx, value=round(float(df[col_name].sum()), 1))
            cell.font = Font(name="Arial", bold=True)

def _nice_scalebar_length(extent_m: float) -> float:
    candidates = [1, 2, 5, 10, 20, 25, 50, 100, 200, 250, 500, 1000, 2000]
    target = extent_m * 0.25
    return min(candidates, key=lambda c: abs(c - target))


def save_cartogram(
    diff: np.ndarray,
    path: str | Path,
    pixel_size_m: float,
    title: str = "Картограмма выемки/насыпи",
    object_masks: list[tuple[int, np.ndarray]] | None = None,
    vmax: float | None = None,
) -> None:
    finite = diff[np.isfinite(diff)]
    if vmax is None:
        vmax = float(np.percentile(np.abs(finite), 99)) if finite.size else 1.0
    vmax = max(vmax, 1e-3)
    norm = TwoSlopeNorm(vmin=-vmax, vcenter=0.0, vmax=vmax)

    fig, ax = plt.subplots(figsize=(9, 8))
    im = ax.imshow(diff, cmap="RdBu", norm=norm, interpolation="nearest")

    if object_masks:
        for obj_id, mask in object_masks:
            for contour in measure.find_contours(mask.astype(float), level=0.5):
                ax.plot(contour[:, 1], contour[:, 0], color="black", linewidth=1.0)
            ys, xs = np.nonzero(mask)
            if len(ys):
                ax.text(
                    xs.mean(), ys.mean(), str(obj_id), color="black", fontsize=9,
                    ha="center", va="center", fontweight="bold",
                    bbox=dict(boxstyle="circle", facecolor="white", alpha=0.75, edgecolor="none"),
                )

    cbar = fig.colorbar(im, ax=ax, shrink=0.85)
    cbar.set_label("Δh, м  (синий — насыпь, красный — выемка)")

    bar_len_m = _nice_scalebar_length(diff.shape[1] * pixel_size_m)
    bar_len_px = bar_len_m / pixel_size_m
    x0, y0 = diff.shape[1] * 0.05, diff.shape[0] * 0.96
    ax.plot([x0, x0 + bar_len_px], [y0, y0], color="black", linewidth=3)
    ax.text(x0 + bar_len_px / 2, y0 - diff.shape[0] * 0.025, f"{bar_len_m:.0f} м", ha="center", fontsize=9)

    ax.set_title(title, fontsize=13)
    ax.set_xlabel("столбцы растра, px")
    ax.set_ylabel("строки растра, px")
    fig.tight_layout()
    fig.savefig(path, dpi=150)
    plt.close(fig)

def export_geojson(
    candidates,
    results_by_id: dict[int, VolumeResult],
    transform: RasterTransform,
    path: str | Path,
) -> None:
    features = []
    for c in candidates:
        contours = measure.find_contours(c.mask.astype(float), level=0.5)
        if not contours:
            continue
        contour = max(contours, key=len)
        ring = [list(transform.pixel_to_geo(row, col)) for row, col in contour]
        if ring[0] != ring[-1]:
            ring.append(ring[0])

        vr = results_by_id.get(c.id)
        props = {"id": c.id, "kind": c.kind}
        if vr is not None:
            props.update(
                {
                    "area_m2": round(vr.area_m2, 1) + 0.0,
                    "cut_volume_m3": round(vr.cut_volume_m3, 1) + 0.0,
                    "fill_volume_m3": round(vr.fill_volume_m3, 1) + 0.0,
                    "net_volume_m3": round(vr.net_volume_m3, 1) + 0.0,
                }
            )
        features.append(
            {"type": "Feature", "geometry": {"type": "Polygon", "coordinates": [ring]}, "properties": props}
        )

    fc = {"type": "FeatureCollection", "features": features}
    Path(path).write_text(json.dumps(fc, ensure_ascii=False, indent=2))
